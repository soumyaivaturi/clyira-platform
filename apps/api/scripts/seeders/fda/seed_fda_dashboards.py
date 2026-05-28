"""
A3. FDA Data Dashboards — bulk XLSX inspection and compliance data.
Output: rag_index/fda_dashboard_inspections.jsonl, rag_index/fda_dashboard_compliance.jsonl
"""
import argparse, json, logging, sys, time, io
from pathlib import Path
from bs4 import BeautifulSoup

SCRIPTS_DIR = Path(__file__).parent.parent.parent
sys.path.insert(0, str(SCRIPTS_DIR.parent))
from scripts.seeders._common import get, get_rag_index, load_existing_compound_keys, append_records, make_id, RAG_INDEX

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

OUTPUT_INSP = RAG_INDEX / "fda_dashboard_inspections.jsonl"
OUTPUT_COMP = RAG_INDEX / "fda_dashboard_compliance.jsonl"

DASHBOARD_URL = "https://datadashboard.fda.gov/oii/cd/"
DIRECT_XLSX_URLS = [
    ("inspections", "https://datadashboard.fda.gov/oii/inspections_data.xlsx"),
    ("inspections", "https://www.fda.gov/media/168914/download"),
    ("compliance", "https://datadashboard.fda.gov/oii/compliance_data.xlsx"),
    ("compliance", "https://www.fda.gov/media/168913/download"),
]


def parse_xlsx_inspections(content: bytes) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(rows[0])]
        records = []
        for row in rows[1:]:
            if not any(row):
                continue
            d = dict(zip(headers, row))
            fei = str(d.get("fei_number", d.get("fei", "")))
            firm = str(d.get("firm_name", d.get("legal_name", d.get("company", ""))))
            insp_date = str(d.get("inspection_end_date", d.get("date_completed", d.get("date", ""))))
            classification = str(d.get("classification", d.get("action_classification", "")))
            city = str(d.get("city", ""))
            state = str(d.get("state_code", d.get("state", "")))
            country = str(d.get("country_iso_2_code", d.get("country", "")))
            product_type = str(d.get("product_type", d.get("program", "")))
            program_area = str(d.get("center_classification_code", d.get("program_area", "")))
            text = (
                f"Inspection: {firm} (FEI {fei}). {city}, {state}, {country}. "
                f"Date: {insp_date}. Classification: {classification}. "
                f"Product type: {product_type}. Program: {program_area}."
            )
            records.append({
                "id": make_id("DASH-INSP", fei, insp_date),
                "source_id": "FDA-DASHBOARD",
                "source_agency": "FDA",
                "source_type": "dashboard_inspection",
                "firm_name": firm,
                "fei_number": fei,
                "city": city,
                "state": state,
                "country": country,
                "inspection_date": insp_date,
                "classification": classification,
                "product_type": product_type,
                "program_area": program_area,
                "text": text,
                "date": insp_date,
                "source_url": DASHBOARD_URL,
            })
        return records
    except Exception as e:
        log.warning(f"Failed to parse inspection XLSX: {e}")
        return []


def parse_xlsx_compliance(content: bytes) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(rows[0])]
        records = []
        for row in rows[1:]:
            if not any(row):
                continue
            d = dict(zip(headers, row))
            action_id = str(d.get("action_id", d.get("id", "")))
            firm = str(d.get("firm_name", d.get("company", "")))
            action_type = str(d.get("action_type", d.get("action", "")))
            date = str(d.get("action_date", d.get("date", "")))
            city = str(d.get("city", ""))
            country = str(d.get("country", ""))
            text = (
                f"Compliance action: {firm}. Action: {action_type}. "
                f"Date: {date}. {city}, {country}."
            )
            records.append({
                "id": make_id("DASH-COMP", action_id or firm, date),
                "source_id": "FDA-DASHBOARD",
                "source_agency": "FDA",
                "source_type": "dashboard_compliance",
                "firm_name": firm,
                "fei_number": str(d.get("fei_number", "")),
                "city": city,
                "state": str(d.get("state", "")),
                "country": country,
                "action_type": action_type,
                "action_date": date,
                "program_area": str(d.get("program_area", "")),
                "text": text,
                "date": date,
                "source_url": DASHBOARD_URL,
            })
        return records
    except Exception as e:
        log.warning(f"Failed to parse compliance XLSX: {e}")
        return []


def main():
    parser = argparse.ArgumentParser(description="Seed FDA Dashboard bulk XLSX data")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    existing_insp = load_existing_compound_keys(OUTPUT_INSP, ["fei_number", "inspection_date"])
    existing_comp = load_existing_compound_keys(OUTPUT_COMP, ["firm_name", "action_date"])

    insp_records = []
    comp_records = []

    # Try each known URL
    for dtype, url in DIRECT_XLSX_URLS:
        log.info(f"Trying {dtype} XLSX: {url}")
        r = get(url, delay=0.5, timeout=120.0)
        if not r or not r.content:
            continue
        content_type = r.headers.get("content-type", "")
        if "html" in content_type:
            continue  # Got HTML instead of XLSX
        if dtype == "inspections" and not insp_records:
            insp_records = parse_xlsx_inspections(r.content)
            log.info(f"  Dashboard inspections: {len(insp_records)} records")
        elif dtype == "compliance" and not comp_records:
            comp_records = parse_xlsx_compliance(r.content)
            log.info(f"  Dashboard compliance: {len(comp_records)} records")

    # If direct URLs fail, try scraping dashboard index
    if not insp_records and not comp_records:
        log.info("Direct XLSX URLs failed — trying dashboard index")
        r = get(DASHBOARD_URL, delay=0.5)
        if r:
            soup = BeautifulSoup(r.text, "html.parser")
            for link in soup.find_all("a", href=True):
                href = link["href"]
                if not href.endswith(".xlsx"):
                    continue
                xlsx_url = href if href.startswith("http") else f"https://datadashboard.fda.gov{href}"
                r2 = get(xlsx_url, delay=0.5, timeout=120.0)
                if not r2:
                    continue
                title = link.get_text(strip=True).lower()
                if "inspection" in title:
                    insp_records = parse_xlsx_inspections(r2.content)
                elif "compliance" in title:
                    comp_records = parse_xlsx_compliance(r2.content)

    new_insp = [r for r in insp_records if (r["fei_number"], r["inspection_date"]) not in existing_insp]
    new_comp = [r for r in comp_records if (r["firm_name"], r["action_date"]) not in existing_comp]

    log.info(f"New inspection records: {len(new_insp)}, compliance: {len(new_comp)}")
    append_records(OUTPUT_INSP, new_insp, args.dry_run, log)
    append_records(OUTPUT_COMP, new_comp, args.dry_run, log)
    log.info(f"FDA Dashboard seeder complete. Inspections: {len(new_insp)}, Compliance: {len(new_comp)}")


if __name__ == "__main__":
    main()
