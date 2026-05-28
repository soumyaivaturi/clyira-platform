"""
B1: MHRA GMP Inspection Deficiency Seeder
==========================================
Downloads MHRA GMP inspection deficiency data from the UK government website.
Primary source: Excel/XLSX spreadsheet published by MHRA.
Output: rag_index/mhra_deficiencies.jsonl

Usage:
    python seed_mhra_deficiencies.py
    python seed_mhra_deficiencies.py --dry-run
"""
import argparse
import json
import logging
import sys
import os
import time
from pathlib import Path

import httpx
from bs4 import BeautifulSoup

SCRIPTS_DIR = Path(__file__).parent.parent.parent
sys.path.insert(0, str(SCRIPTS_DIR.parent))
from scripts.seeders._common import get, get_rag_index, load_existing_compound_keys, append_records, pdf_to_text, make_id, RAG_INDEX

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

OUTPUT_FILE = "mhra_deficiencies.jsonl"

# Primary and fallback URLs for the MHRA GMP deficiencies spreadsheet
PRIMARY_EXCEL_URL = (
    "https://assets.publishing.service.gov.uk/government/uploads/system/uploads/"
    "attachment_data/file/gmp-deficiencies.xlsx"
)
STATS_PAGE_URL = (
    "https://www.gov.uk/government/statistics/good-manufacturing-practice-inspection-deficiencies"
)
ALTERNATE_STATS_URL = (
    "https://www.gov.uk/government/statistics?keywords=gmp+deficiencies&"
    "topics%5B%5D=all&departments%5B%5D=medicines-and-healthcare-products-regulatory-agency"
)


def find_excel_url_from_page(html: str) -> str | None:
    """Parse the GOV.UK stats page to find the current Excel download link."""
    soup = BeautifulSoup(html, "lxml")
    # GOV.UK attachment links are usually in .attachment-details or .gem-c-attachment
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.endswith(".xlsx") or href.endswith(".xls"):
            if "gmp" in href.lower() or "deficien" in href.lower() or "deficien" in a.get_text(strip=True).lower():
                if href.startswith("http"):
                    return href
                return "https://assets.publishing.service.gov.uk" + href
    # Broader search for any xlsx on the page
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.endswith(".xlsx") or href.endswith(".xls"):
            if href.startswith("http"):
                return href
            return "https://assets.publishing.service.gov.uk" + href
    return None


def fetch_excel_bytes() -> tuple[bytes | None, str]:
    """Try multiple URLs to get the Excel file. Returns (bytes, url_used)."""
    # Step 1: try the static asset URL directly
    log.info("Trying direct asset URL for MHRA deficiencies Excel…")
    resp = get(PRIMARY_EXCEL_URL, delay=1.0, timeout=60.0)
    if resp and resp.status_code == 200 and len(resp.content) > 1000:
        log.info(f"  Downloaded {len(resp.content):,} bytes from primary URL")
        return resp.content, PRIMARY_EXCEL_URL

    # Step 2: scrape the statistics page to find current link
    log.info("Primary URL failed, scraping statistics page…")
    resp = get(STATS_PAGE_URL, delay=1.0, timeout=30.0)
    if resp:
        excel_url = find_excel_url_from_page(resp.text)
        if excel_url:
            log.info(f"  Found Excel URL from stats page: {excel_url}")
            time.sleep(1.0)
            r2 = get(excel_url, delay=1.0, timeout=60.0)
            if r2 and len(r2.content) > 1000:
                return r2.content, excel_url

    # Step 3: try alternate search page
    log.info("Trying alternate GOV.UK search URL…")
    resp = get(ALTERNATE_STATS_URL, delay=1.0, timeout=30.0)
    if resp:
        excel_url = find_excel_url_from_page(resp.text)
        if excel_url:
            time.sleep(1.0)
            r3 = get(excel_url, delay=1.0, timeout=60.0)
            if r3 and len(r3.content) > 1000:
                return r3.content, excel_url

    log.warning("All attempts to download MHRA deficiencies Excel failed — returning empty")
    return None, ""


def parse_excel(content: bytes, source_url: str) -> list[dict]:
    """Parse the MHRA deficiencies Excel into records."""
    try:
        import openpyxl
        import io
    except ImportError:
        log.error("openpyxl not installed; run: pip install openpyxl")
        return []

    records = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        log.error(f"Failed to open workbook: {e}")
        return []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect header row (first row with non-empty strings)
        header_idx = 0
        headers = []
        for i, row in enumerate(rows):
            non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if len(non_empty) >= 3:
                headers = [str(c).strip().lower() if c else "" for c in row]
                header_idx = i
                break

        if not headers:
            continue

        log.info(f"  Sheet '{sheet_name}': {len(rows) - header_idx - 1} data rows, headers: {headers[:8]}")

        # Map column positions
        col = {}
        for i, h in enumerate(headers):
            if any(k in h for k in ["deficien", "observation", "finding", "text", "description"]):
                col.setdefault("deficiency_text", i)
            elif any(k in h for k in ["year", "inspection year", "date"]):
                col.setdefault("year", i)
            elif any(k in h for k in ["class", "classif", "critical", "major", "other"]):
                col.setdefault("classification", i)
            elif any(k in h for k in ["eu gmp", "guideline", "reference", "chapter ref", "part"]):
                col.setdefault("eu_gmp_reference", i)
            elif "chapter" in h:
                col.setdefault("chapter", i)
            elif "annex" in h:
                col.setdefault("annex", i)

        for row in rows[header_idx + 1:]:
            try:
                def cell(key):
                    idx = col.get(key)
                    return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ""

                deficiency_text = cell("deficiency_text")
                if not deficiency_text or deficiency_text.lower() in ("none", "n/a", ""):
                    continue

                year = cell("year")
                # Extract 4-digit year if full date
                import re
                year_match = re.search(r"\b(20\d{2}|19\d{2})\b", year)
                year_clean = year_match.group(1) if year_match else year[:4]

                classification = cell("classification")
                eu_gmp_ref = cell("eu_gmp_reference")
                chapter = cell("chapter")
                annex = cell("annex")

                text = (
                    f"MHRA GMP Inspection Deficiency ({year_clean}). "
                    f"Classification: {classification}. "
                    f"EU GMP Reference: {eu_gmp_ref}. "
                    f"Chapter: {chapter}. Annex: {annex}. "
                    f"Deficiency: {deficiency_text}"
                )

                records.append({
                    "id": make_id("MHRA-DEF", deficiency_text[:100], year_clean),
                    "source_id": "MHRA-DEF",
                    "source_agency": "MHRA",
                    "source_type": "gmp_inspection_deficiency",
                    "year": year_clean,
                    "deficiency_text": deficiency_text,
                    "classification": classification,
                    "eu_gmp_reference": eu_gmp_ref,
                    "chapter": chapter,
                    "annex": annex,
                    "text": text,
                    "date": f"{year_clean}-01-01" if year_clean.isdigit() and len(year_clean) == 4 else year_clean,
                    "source_url": source_url,
                })
            except Exception as e:
                log.debug(f"Skipping row: {e}")
                continue

    wb.close()
    log.info(f"Parsed {len(records)} deficiency records from Excel")
    return records


def main():
    parser = argparse.ArgumentParser(description="Seed MHRA GMP deficiencies into RAG index")
    parser.add_argument("--dry-run", action="store_true", help="Print records without writing")
    args = parser.parse_args()

    out_path = get_rag_index() / OUTPUT_FILE
    existing = load_existing_compound_keys(out_path, ["deficiency_text", "year"])
    log.info(f"Existing records in {OUTPUT_FILE}: {len(existing)}")

    content, url_used = fetch_excel_bytes()
    if not content:
        log.warning("No Excel data retrieved — exiting with 0 new records")
        return

    all_records = parse_excel(content, url_used)

    new_records = []
    for r in all_records:
        key = (r["deficiency_text"][:100], r["year"])
        if key not in existing:
            new_records.append(r)
            existing.add(key)

    log.info(f"New records after dedup: {len(new_records)}")
    written = append_records(out_path, new_records, args.dry_run, log)
    log.info(f"{'Would write' if args.dry_run else 'Wrote'} {written} records to {out_path}")


if __name__ == "__main__":
    main()
